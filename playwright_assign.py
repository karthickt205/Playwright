import argparse
import json
import random
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


WHATSAPP_URL = "https://web.whatsapp.com/"
DEFAULT_MESSAGE = "Hello {name}"
ACTION_TIMEOUT = 20_000


def random_delay(page: Page) -> None:
	page.wait_for_timeout(random.randint(2_000, 5_000))


def mask_phone(phone: str) -> str:
	return "*" * max(0, len(phone) - 4) + phone[-4:]


def read_contacts(path: Path) -> list[dict[str, str]]:
	workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
	sheet = workbook.active
	rows = sheet.iter_rows(values_only=True)
	headers = [str(value or "").strip().lower() for value in next(rows, ())]
	required = {"name", "phone", "message"}
	if not required.issubset(headers):
		raise ValueError("contacts.xlsx must contain Name, Phone, and Message columns")

	positions = {header: index for index, header in enumerate(headers)}
	contacts = []
	for row in rows:
		name = str(row[positions["name"]] or "").strip()
		phone = re.sub(r"\D", "", str(row[positions["phone"]] or ""))
		message = str(row[positions["message"]] or "").strip()
		if name or phone:
			contacts.append({"name": name, "phone": phone, "message": message})
	return contacts


def wait_for_login(page: Page) -> None:
	page.goto(WHATSAPP_URL, wait_until="domcontentloaded")
	try:
		page.wait_for_selector('div[aria-label="Chat list"], [data-testid="chat-list"]', timeout=8_000)
		return
	except PlaywrightTimeoutError:
		print("Scan the WhatsApp Web QR code in the browser, then press Enter here.")
		input()
	page.wait_for_selector('div[aria-label="Chat list"], [data-testid="chat-list"]', timeout=120_000)


def search_contact(page: Page, contact: dict[str, str]) -> None:
	search = page.locator(
		'div[contenteditable="true"][data-tab="3"], '
		'input[placeholder*="Search"], div[role="textbox"][aria-label*="Search"]'
	).first
	search.wait_for(state="visible", timeout=ACTION_TIMEOUT)
	search.click()
	search.fill(contact["phone"] or contact["name"])
	random_delay(page)
	result = page.locator(
		'[data-testid="cell-frame-container"], '
		'div[role="gridcell"], span[title]'
	).filter(has_text=re.compile(re.escape(contact["phone"] or contact["name"]), re.I)).first
	result.wait_for(state="visible", timeout=ACTION_TIMEOUT)
	result.click()


def send_message(page: Page, message: str) -> None:
	message_box = page.locator(
		'footer div[contenteditable="true"], '
		'div[contenteditable="true"][data-tab="10"], '
		'[aria-label="Type a message"]'
	).first
	message_box.wait_for(state="visible", timeout=ACTION_TIMEOUT)
	message_box.click()
	message_box.fill("")
	page.keyboard.insert_text(message)
	random_delay(page)
	send_button = page.locator(
		'button[aria-label="Send"], '
		'[data-testid="send"], '
		'span[data-icon="send"]'
	).first
	if send_button.is_visible():
		send_button.click()
	else:
		page.keyboard.press("Enter")
	page.wait_for_selector('[data-testid="msg-container"]', state="visible", timeout=ACTION_TIMEOUT)
	page.wait_for_timeout(2_000)
	sent_message = page.locator('[data-testid="msg-container"]').filter(
		has_text=re.compile(re.escape(message), re.I)
	).last
	sent_message.wait_for(state="visible", timeout=ACTION_TIMEOUT)


def extract_last_messages(page: Page, limit: int = 3) -> list[str]:
	message_nodes = page.locator(
		'[data-testid="msg-container"] span.selectable-text, '
		'[data-testid="msg-container"] [dir="auto"]'
	)
	count = message_nodes.count()
	messages = []
	for index in range(max(0, count - limit), count):
		text = message_nodes.nth(index).inner_text().strip()
		if text and (not messages or messages[-1] != text):
			messages.append(text)
	return messages[-limit:]


def process_contact(page: Page, contact: dict[str, str], screenshot_dir: Path) -> dict:
	started = datetime.now().isoformat(timespec="seconds")
	message = (contact["message"] or DEFAULT_MESSAGE).replace("{name}", contact["name"])
	result = {
		"name": contact["name"],
		"phone": mask_phone(contact["phone"]),
		"message": message,
		"sent": False,
		"last_messages": [],
		"screenshot": "",
		"error": "",
		"processed_at": started,
	}
	try:
		search_contact(page, contact)
		random_delay(page)
		send_message(page, message)
		page.wait_for_selector('[data-testid="msg-container"]', state="visible", timeout=ACTION_TIMEOUT)
		safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", contact["name"] or "contact")
		safe_phone = re.sub(r"[^0-9]+", "", contact["phone"])
		screenshot_path = (screenshot_dir / f"{safe_name}.png").resolve()
		screenshot_path.parent.mkdir(parents=True, exist_ok=True)
		page.screenshot(path=str(screenshot_path), full_page=False)
		if not screenshot_path.is_file():
			raise IOError(f"Screenshot was not created: {screenshot_path}")
		result["screenshot"] = str(screenshot_path)
		result["sent"] = True
		result["last_messages"] = extract_last_messages(page)
		print(f"Screenshot saved: {screenshot_path}")
	except Exception as error:
		result["error"] = f"{type(error).__name__}: {error}"
	return result


def save_reports(results: list[dict], output_dir: Path) -> None:
	output_dir.mkdir(parents=True, exist_ok=True)
	timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
	json_path = output_dir / f"whatsapp_report_{timestamp}.json"
	excel_path = output_dir / f"whatsapp_report_{timestamp}.xlsx"
	json_path.write_text(json.dumps(results, indent=2, ensure_ascii=True), encoding="utf-8")

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = "Report"
	headers = ["Name", "Phone", "Message", "Sent", "Last 3 Messages", "Screenshot", "Error", "Processed At"]
	sheet.append(headers)
	for result in results:
		sheet.append([
			result["name"], result["phone"], result["message"], result["sent"],
			"\n".join(result["last_messages"]), result["screenshot"], result["error"], result["processed_at"],
		])
	sheet.freeze_panes = "A2"
	sheet.auto_filter.ref = sheet.dimensions
	for column in sheet.columns:
		sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(column[0].value or "")) + 2, 12), 40)
	workbook.save(excel_path)
	print(f"Reports saved to {json_path} and {excel_path}")


def main() -> None:
	parser = argparse.ArgumentParser(description="Send personalized WhatsApp Web messages from Excel.")
	parser.add_argument("--contacts", type=Path, default=Path("contact.xlsx"))
	parser.add_argument("--output", type=Path, default=Path("whatsapp_reports"))
	parser.add_argument("--browser-data", type=Path, default=Path(".whatsapp_browser_data"))
	args = parser.parse_args()
	contacts_path = args.contacts
	if not contacts_path.is_absolute() and not contacts_path.exists():
		contacts_path = Path(__file__).resolve().parent / contacts_path
	if not contacts_path.is_file():
		raise FileNotFoundError(f"Contacts workbook not found: {contacts_path.resolve()}")
	contacts = read_contacts(contacts_path)
	screenshot_dir = args.output / "screenshots"
	screenshot_dir.mkdir(parents=True, exist_ok=True)

	results = []
	with sync_playwright() as playwright:
		context = playwright.chromium.launch_persistent_context(
			user_data_dir=str(args.browser_data), headless=False
		)
		page = context.pages[0] if context.pages else context.new_page()
		try:
			wait_for_login(page)
			for contact in contacts:
				results.append(process_contact(page, contact, screenshot_dir))
				random_delay(page)
		finally:
			context.close()
	save_reports(results, args.output)


if __name__ == "__main__":
	main()
